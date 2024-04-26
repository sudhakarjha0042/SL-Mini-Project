import openpyxl

def send_seat_numbers():
    wb_branch = openpyxl.load_workbook(r'Exam-Seat-Arrangement-System-main\ESA-Backend\updatedExcels\16-04-2024_FN.xlsx')

    # Iterate over all sheets starting from the third one
    for sheet_name in wb_branch.sheetnames[2:]:
        sheet = wb_branch[sheet_name]

        # Extract the data into a list of dictionaries
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        # Skip the header row
        for row in data[1:]:
            seat_no = row[1]
            email = row[2]
            message = f'Your seat number is {seat_no}. Your room number is {sheet_name}.'
            # send_email(email, message)
            print(message)
            print("\n")

# Call the function
send_seat_numbers()